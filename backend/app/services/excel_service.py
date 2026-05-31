import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import calendar
import io
import logging
from sqlalchemy.orm import Session
from app.models.models import (
    Employee, AttendanceSummary, AttendanceLog,
    AttendanceStatus, AttendanceType, Department
)

logger = logging.getLogger(__name__)

# ── colour palette ────────────────────────────────────────────
TEAL   = "00D4AA"
DARK   = "0D1117"
BLUE   = "0EA5E9"
GREEN  = "3FB950"
RED    = "F85149"
YELLOW = "D29922"
GREY   = "8B949E"
WHITE  = "FFFFFF"
LIGHT  = "F6F8FA"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center")


class ExcelService:

    # ── 1. Daily Attendance Report ─────────────────────────────
    def generate_daily_report(self, db: Session, report_date: date) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Attendance"
        ws.sheet_view.showGridLines = False

        # Title row
        ws.merge_cells("A1:H1")
        ws["A1"] = "DAILY ATTENDANCE REPORT"
        ws["A1"].font      = _font(bold=True, color=WHITE, size=14)
        ws["A1"].fill      = _fill(DARK)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 35

        # Date row
        ws.merge_cells("A2:H2")
        ws["A2"] = "Date: " + report_date.strftime("%B %d, %Y") + \
                   "   |   Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["A2"].font      = _font(color=WHITE, size=10)
        ws["A2"].fill      = _fill(TEAL)
        ws["A2"].alignment = _center()
        ws.row_dimensions[2].height = 22

        # Header
        headers = ["#", "Employee ID", "Employee Name", "Department",
                   "Status", "First In", "Last Out", "Work Hours"]
        col_w   = [5, 14, 22, 18, 12, 12, 12, 12]
        for i, (h, w) in enumerate(zip(headers, col_w), 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font      = _font(bold=True, color=WHITE, size=10)
            cell.fill      = _fill("161B22")
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[3].height = 22

        # Data
        summaries = db.query(AttendanceSummary).filter(
            AttendanceSummary.date == report_date
        ).join(Employee).order_by(Employee.first_name).all()

        status_colors = {
            "PRESENT":  GREEN,
            "LATE":     YELLOW,
            "ABSENT":   RED,
            "HALF_DAY": BLUE,
        }
        present = late = absent = 0

        for idx, s in enumerate(summaries, 1):
            emp   = s.employee
            row   = idx + 3
            bg    = LIGHT if idx % 2 == 0 else WHITE
            sname = s.status.value if s.status else "ABSENT"

            if sname == "PRESENT": present += 1
            elif sname == "LATE":  present += 1; late += 1
            else:                  absent  += 1

            values = [
                idx,
                emp.employee_id,
                emp.full_name,
                emp.department.name if emp.department else "-",
                sname,
                s.first_in.strftime("%H:%M:%S")  if s.first_in  else "-",
                s.last_out.strftime("%H:%M:%S") if s.last_out else "-",
                str(round(s.work_hours, 2)) + " hrs" if s.work_hours else "-",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = _fill(bg)
                cell.border    = _border()
                cell.alignment = _center() if col != 3 else _left()
                cell.font      = Font(size=10, color="333333")
                if col == 5:
                    sc = status_colors.get(sname, GREY)
                    cell.font = Font(bold=True, color=sc, size=10)
            ws.row_dimensions[row].height = 18

        # Summary box
        last_row = len(summaries) + 5
        ws.merge_cells("A" + str(last_row) + ":H" + str(last_row))
        ws["A" + str(last_row)] = (
            "SUMMARY   |   Total: " + str(len(summaries)) +
            "   Present: " + str(present) +
            "   Late: " + str(late) +
            "   Absent: " + str(absent) +
            "   Attendance Rate: " +
            str(round(present / len(summaries) * 100 if summaries else 0, 1)) + "%"
        )
        ws["A" + str(last_row)].font      = _font(bold=True, color=WHITE, size=10)
        ws["A" + str(last_row)].fill      = _fill(TEAL)
        ws["A" + str(last_row)].alignment = _center()
        ws.row_dimensions[last_row].height = 24

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info("Daily report generated for: " + str(report_date))
        return buf.getvalue()

    # ── 2. Monthly Attendance Report ───────────────────────────
    def generate_monthly_report(self, db: Session, year: int, month: int) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        ws.sheet_view.showGridLines = False

        month_name  = datetime(year, month, 1).strftime("%B %Y")
        total_days  = calendar.monthrange(year, month)[1]
        first_day   = date(year, month, 1)
        last_day    = date(year, month, total_days)

        # Title
        total_cols = total_days + 6
        ws.merge_cells("A1:" + get_column_letter(total_cols) + "1")
        ws["A1"] = "MONTHLY ATTENDANCE REPORT — " + month_name.upper()
        ws["A1"].font      = _font(bold=True, size=14)
        ws["A1"].fill      = _fill(DARK)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 35

        # Fixed headers
        fixed = ["#", "Emp ID", "Name", "Dept"]
        for i, h in enumerate(fixed, 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font      = _font(bold=True, size=9)
            cell.fill      = _fill("161B22")
            cell.alignment = _center()
            cell.border    = _border()

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 14

        # Day headers
        for d in range(1, total_days + 1):
            col  = d + 4
            cell = ws.cell(row=2, column=col, value=d)
            cell.font      = _font(bold=True, size=8)
            cell.fill      = _fill(BLUE)
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col)].width = 4

        # Summary headers
        for i, h in enumerate(["Present", "Late", "Absent", "Hours"], 1):
            col  = total_days + 4 + i
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _font(bold=True, size=9)
            cell.fill      = _fill(TEAL)
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col)].width = 9

        ws.row_dimensions[2].height = 22

        # Employee data
        employees = db.query(Employee).filter(
            Employee.is_active == True
        ).order_by(Employee.first_name).all()

        for idx, emp in enumerate(employees, 1):
            row = idx + 2
            bg  = LIGHT if idx % 2 == 0 else WHITE

            summaries = {
                s.date: s for s in db.query(AttendanceSummary).filter(
                    AttendanceSummary.employee_id == emp.id,
                    AttendanceSummary.date >= first_day,
                    AttendanceSummary.date <= last_day
                ).all()
            }

            for col, val in enumerate([idx, emp.employee_id, emp.full_name,
                                        emp.department.name if emp.department else "-"], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = _fill(bg)
                cell.border    = _border()
                cell.alignment = _left() if col == 3 else _center()
                cell.font      = Font(size=9, color="333333")

            present = late = absent = total_hours = 0

            for d in range(1, total_days + 1):
                col       = d + 4
                day_date  = date(year, month, d)
                summary   = summaries.get(day_date)
                cell      = ws.cell(row=row, column=col)
                cell.border    = _border()
                cell.alignment = _center()
                cell.font      = Font(size=8, bold=True)

                if summary:
                    st = summary.status.value if summary.status else "A"
                    if st == "PRESENT":
                        cell.value = "P"
                        cell.font  = Font(size=8, bold=True, color=GREEN)
                        present += 1
                        total_hours += summary.work_hours or 0
                    elif st == "LATE":
                        cell.value = "L"
                        cell.font  = Font(size=8, bold=True, color=YELLOW)
                        present += 1
                        late    += 1
                        total_hours += summary.work_hours or 0
                    elif st == "HALF_DAY":
                        cell.value = "H"
                        cell.font  = Font(size=8, bold=True, color=BLUE)
                        present += 1
                        total_hours += summary.work_hours or 0
                    else:
                        cell.value = "A"
                        cell.font  = Font(size=8, bold=True, color=RED)
                        absent += 1
                else:
                    cell.value = "-"
                    cell.font  = Font(size=8, color=GREY)

            for i, val in enumerate([present, late, absent,
                                      str(round(total_hours, 1)) + "h"], 1):
                col  = total_days + 4 + i
                cell = ws.cell(row=row, column=col, value=val)
                cell.border    = _border()
                cell.alignment = _center()
                cell.font      = Font(size=9, bold=True,
                                      color=GREEN if i == 1 else
                                            YELLOW if i == 2 else
                                            RED    if i == 3 else BLUE)
                cell.fill = _fill(bg)

            ws.row_dimensions[row].height = 16

        # Legend
        legend_row = len(employees) + 4
        ws.merge_cells("A" + str(legend_row) + ":D" + str(legend_row))
        ws["A" + str(legend_row)] = "Legend:"
        ws["A" + str(legend_row)].font = Font(bold=True, size=9)

        for i, (sym, label, color) in enumerate([
            ("P", "Present", GREEN), ("L", "Late", YELLOW),
            ("A", "Absent", RED),   ("H", "Half Day", BLUE)
        ], 1):
            col  = i
            cell = ws.cell(row=legend_row + 1, column=col,
                           value=sym + " = " + label)
            cell.font      = Font(bold=True, color=color, size=9)
            cell.alignment = _center()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info("Monthly report generated: " + str(month) + "/" + str(year))
        return buf.getvalue()

    # ── 3. Salary Sheet ────────────────────────────────────────
    def generate_salary_sheet(self, db: Session, year: int, month: int,
                               basic_salary_map: dict = None) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Sheet"
        ws.sheet_view.showGridLines = False

        month_name = datetime(year, month, 1).strftime("%B %Y")
        total_days = calendar.monthrange(year, month)[1]
        first_day  = date(year, month, 1)
        last_day   = date(year, month, total_days)

        # Title
        ws.merge_cells("A1:M1")
        ws["A1"] = "SALARY SHEET — " + month_name.upper()
        ws["A1"].font      = _font(bold=True, size=14)
        ws["A1"].fill      = _fill(DARK)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:M2")
        ws["A2"] = "Confidential — HR Department"
        ws["A2"].font      = _font(size=9, color="CCCCCC")
        ws["A2"].fill      = _fill("161B22")
        ws["A2"].alignment = _center()

        # Headers
        headers = [
            "#", "Emp ID", "Name", "Department",
            "Working Days", "Present", "Absent", "Late",
            "Basic Salary", "Overtime", "Deduction", "Net Salary", "Status"
        ]
        widths = [4, 12, 22, 16, 13, 10, 10, 8, 13, 12, 12, 13, 10]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font      = _font(bold=True, size=10)
            cell.fill      = _fill(TEAL)
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[3].height = 24

        employees = db.query(Employee).filter(
            Employee.is_active == True
        ).order_by(Employee.first_name).all()

        total_net = 0

        for idx, emp in enumerate(employees, 1):
            row = idx + 3
            bg  = LIGHT if idx % 2 == 0 else WHITE

            summaries = db.query(AttendanceSummary).filter(
                AttendanceSummary.employee_id == emp.id,
                AttendanceSummary.date >= first_day,
                AttendanceSummary.date <= last_day
            ).all()

            present     = sum(1 for s in summaries if s.status in [
                AttendanceStatus.PRESENT, AttendanceStatus.LATE])
            absent      = sum(1 for s in summaries if s.status == AttendanceStatus.ABSENT)
            late        = sum(1 for s in summaries if s.is_late)
            total_hours = sum(s.work_hours or 0 for s in summaries)

            basic       = (basic_salary_map or {}).get(emp.employee_id, 12000)
            per_day     = basic / total_days
            overtime    = max(0, total_hours - (present * 8)) * (per_day / 8)
            deduction   = absent * per_day + (late * per_day * 0.1)
            net         = max(0, basic + overtime - deduction)
            total_net  += net

            values = [
                idx, emp.employee_id, emp.full_name,
                emp.department.name if emp.department else "-",
                total_days, present, absent, late,
                basic, round(overtime, 2),
                round(deduction, 2), round(net, 2),
                "PAID" if net > 0 else "REVIEW"
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = _fill(bg)
                cell.border    = _border()
                cell.alignment = _left() if col == 3 else _center()
                cell.font      = Font(size=10, color="333333")

                if col == 12:
                    cell.font      = Font(bold=True, color=GREEN, size=10)
                    cell.number_format = "#,##0.00"
                elif col in [9, 10, 11]:
                    cell.number_format = "#,##0.00"
                elif col == 13:
                    color = GREEN if val == "PAID" else YELLOW
                    cell.font = Font(bold=True, color=color, size=10)

            ws.row_dimensions[row].height = 18

        # Total row
        total_row = len(employees) + 5
        ws.merge_cells("A" + str(total_row) + ":K" + str(total_row))
        ws["A" + str(total_row)] = "TOTAL NET SALARY PAYABLE"
        ws["A" + str(total_row)].font      = _font(bold=True, size=11)
        ws["A" + str(total_row)].fill      = _fill(DARK)
        ws["A" + str(total_row)].alignment = _center()

        total_cell = ws.cell(row=total_row, column=12,
                             value=round(total_net, 2))
        total_cell.font          = _font(bold=True, color=TEAL, size=12)
        total_cell.fill          = _fill(DARK)
        total_cell.alignment     = _center()
        total_cell.number_format = "#,##0.00"
        ws.row_dimensions[total_row].height = 28

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info("Salary sheet generated: " + str(month) + "/" + str(year))
        return buf.getvalue()


excel_service = ExcelService()
